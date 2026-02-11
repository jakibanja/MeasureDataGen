import pandas as pd
import yaml
import os
import argparse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Protection

class SmartTemplateGenerator:
    """
    Generates 'Smart' Excel templates for HEDIS test cases.
    Features:
    - Standard Format compliance
    - Drop-down lists (Data Validation) for Event Names
    - Conditional Formatting (basic)
    - Instructions and comments
    """
    
    def __init__(self, measure_name, config_path=None):
        self.measure_name = measure_name.upper()
        if not config_path:
            config_path = f'config/{self.measure_name}.yaml'
        self.config_path = config_path
        self.config = {}
        self.load_config()

    def load_config(self):
        """Load measure configuration."""
        if os.path.exists(self.config_path):
            with open(self.config_path, 'r') as f:
                self.config = yaml.safe_load(f)
            print(f"✅ Loaded config for {self.measure_name}")
        else:
            print(f"⚠️ Config not found at {self.config_path}. Template will be generic.")

    def get_valid_events(self):
        """Extract valid event names from config."""
        events = []
        if 'rules' in self.config and 'clinical_events' in self.config['rules']:
            # Numerator components
            for comp in self.config['rules']['clinical_events'].get('numerator_components', []):
                events.append(comp['name'])
            # Denominator components
            for comp in self.config['rules']['clinical_events'].get('denominator_components', []):
                events.append(comp['name'])
        return list(set(events)) # Unique

    def get_valid_exclusions(self):
        """Extract valid exclusion names from config."""
        exclusions = []
        if 'rules' in self.config and 'exclusions' in self.config['rules']:
            for excl in self.config['rules']['exclusions']:
                exclusions.append(excl['name'])
        return list(set(exclusions))

    def generate_template(self, output_path=None):
        """Create the Excel template file."""
        if not output_path:
            output_path = f'data/{self.measure_name}_TestCase_SMART.xlsx'
            
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.measure_name} Test Cases"
        
        # 1. Define Headers (Standard Format)
        headers = [
            'MEMBER_ID', 'AGE', 'GENDER', 'PRODUCT_LINE',
            'ENROLLMENT_1_START', 'ENROLLMENT_1_END',
            'VISIT_1_DATE', 'VISIT_1_TYPE',
            'EVENT_1_NAME', 'EVENT_1_VALUE', 'EVENT_1_DATE',
            'EVENT_2_NAME', 'EVENT_2_VALUE', 'EVENT_2_DATE',
            'EXCLUSION_1_NAME', 'EXCLUSION_1_VALUE', 'EXCLUSION_1_DATE',
            'EXPECTED_RESULT', 'NOTES'
        ]
        
        ws.append(headers)
        
        # 2. Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # 3. Data Validation (The "Smart" Part)
        valid_events = self.get_valid_events()
        valid_exclusions = self.get_valid_exclusions()
        
        # Event Name Validation (Columns I and L - 9 and 12)
        if valid_events:
            dv_event = DataValidation(type="list", formula1=f'"{",".join(valid_events)}"', allow_blank=True)
            dv_event.error = 'Invalid Event Name'
            dv_event.errorTitle = 'Error'
            dv_event.prompt = 'Select a valid Clinical Event from the list'
            dv_event.promptTitle = 'Select Event'
            
            ws.add_data_validation(dv_event)
            # Apply to rows 2-1000 for Event Columns
            dv_event.add(f'I2:I1000') 
            dv_event.add(f'L2:L1000')
            
        # Exclusion Name Validation (Column O - 15)
        if valid_exclusions:
            dv_excl = DataValidation(type="list", formula1=f'"{",".join(valid_exclusions)}"', allow_blank=True)
            ws.add_data_validation(dv_excl)
            dv_excl.add(f'O2:O1000')

        # Gender Validation
        dv_gender = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
        ws.add_data_validation(dv_gender)
        dv_gender.add('C2:C1000')

        # Product Line Validation
        dv_pl = DataValidation(type="list", formula1='"Commercial,Medicare,Medicaid,Exchange"', allow_blank=True)
        ws.add_data_validation(dv_pl)
        dv_pl.add('D2:D1000')

        # 4. Add Sample Rows (to avoid "empty" feel)
        self.add_sample_rows(ws, valid_events, valid_exclusions)

        # 5. Instructions Row (Optional, maybe in a separate sheet)
        # For now, just setting column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        print(f"✅ Smart Template generated: {output_path}")
        print(f"   - Events linked: {valid_events}")
        print(f"   - Exclusions linked: {valid_exclusions}")

    def add_sample_rows(self, ws, valid_events, valid_exclusions):
        """Add 2-3 sample rows to guide the user."""
        event_1 = valid_events[0] if valid_events else "Sample Screening"
        excl_1 = valid_exclusions[0] if valid_exclusions else "Hospice"
        
        samples = [
            [
                f'{self.measure_name}_MEMBER_01', 65, 'M', 'Medicare',
                '2026-01-01', '2026-12-31', 
                '2026-03-15', 'Outpatient', 
                event_1, 'Compliant', '2026-05-20',
                '', '', '',
                '', '', '',
                '1', 'Sample Compliant Case'
            ],
            [
                f'{self.measure_name}_MEMBER_02', 45, 'F', 'Commercial',
                '2026-01-01', '2026-12-31', 
                '2026-06-10', 'Office Visit', 
                '', '', '',
                '', '', '',
                excl_1, 'Excluded', '2026-08-15',
                '0', 'Sample Exclusion Case'
            ]
        ]
        
        for row in samples:
            ws.append(row)
            
        print(f"   - Added {len(samples)} sample rows")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Generate Smart Excel Template')
    parser.add_argument('measure', help='Measure Name (e.g. PSA)')
    args = parser.parse_args()
    
    generator = SmartTemplateGenerator(args.measure)
    generator.generate_template()
